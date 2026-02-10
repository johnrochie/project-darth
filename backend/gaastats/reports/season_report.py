"""
Season Report Generation
Excel report for entire club season
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.conf import settings
from ..models import Match, Player, MatchEvent

REPORTS_DIR = settings.MEDIA_ROOT / 'reports'
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def generate_season_report_excel(club) -> str:
    """Generate Excel report for entire season"""

    # Get all matches
    matches = Match.objects.filter(club=club).order_by('-date')

    # Calculate season-wide stats
    club_stats = {
        'matches': matches.count(),
        'goals': 0,
        'point_1': 0,
        'point_2': 0,
        'shot_taken': 0,
        'shot_on_target': 0,
        'tackle_won': 0,
    }

    for match in matches:
        club_stats['goals'] += match.club_goals
        club_stats['point_1'] += match.club_1point
        club_stats['point_2'] += match.club_2point

        # Get events for shots/tackles
        events = MatchEvent.objects.filter(match=match)
        for event in events:
            if event.event_type in ['shot_on_target', 'shot_wide', 'shot_saved']:
                club_stats['shot_taken'] += 1
                if event.event_type == 'shot_on_target':
                    club_stats['shot_on_target'] += 1
            elif event.event_type == 'tackle_won':
                club_stats['tackle_won'] += 1

    total_score = club_stats['goals'] * 3 + club_stats['point_1'] + club_stats['point_2']
    accuracy = (club_stats['shot_on_target'] / club_stats['shot_taken'] * 100) if club_stats['shot_taken'] > 0 else 0

    # Player stats
    player_stats = {}

    for event in MatchEvent.objects.filter(match__club=club).select_related('player', 'match'):
        if not event.player:
            continue

        if event.player_id not in player_stats:
            player_stats[event.player_id] = {
                'number': event.player.number,
                'name': event.player.name,
                'goals': 0,
                'point_1': 0,
                'point_2': 0,
                'shots_taken': 0,
                'shots_on_target': 0,
                'tackles_won': 0,
                'matches': set(),
            }

        stats = player_stats[event.player_id]

        if event.event_type == 'score_goal':
            stats['goals'] += 1
        elif event.event_type == 'score_1point':
            stats['point_1'] += 1
        elif event.event_type == 'score_2point':
            stats['point_2'] += 1
        elif event.event_type in ['shot_on_target', 'shot_wide', 'shot_saved']:
            stats['shots_taken'] += 1
            if event.event_type == 'shot_on_target':
                stats['shots_on_target'] += 1
        elif event.event_type == 'tackle_won':
            stats['tackles_won'] += 1

        stats['matches'].add(event.match_id)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"{club.name} - Season Report"

    # Styles
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center")
    left_alignment = Alignment(horizontal="left")

    # Club Info
    row = 1
    ws[f'A{row}'] = f"{club.name} - Season Report"
    ws[f'A{row}'].font = Font(bold=True, size=18)

    row += 2
    ws[f'A{row}'] = 'Total Matches:'
    ws[f'B{row}'] = club_stats['matches']
    ws[f'A{row}'].font = Font(bold=True)

    row += 2

    # Team Stats
    ws[f'A{row}'] = 'Season Statistics'
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    headers = ['Goals', '1-Points', '2-Points', 'Total', 'Shots', 'Accuracy', 'Tackles']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    row += 1
    team_values = [
        club_stats['goals'],
        club_stats['point_1'],
        club_stats['point_2'],
        total_score,
        club_stats['shot_taken'],
        f"{accuracy:.1f}%",
        club_stats['tackle_won'],
    ]

    for col, value in enumerate(team_values, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.alignment = center_alignment

    row += 2

    # Player Stats Header
    ws[f'A{row}'] = 'Player Statistics'
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    player_headers = ['#', 'Player', 'Goals', '1-Points', '2-Points', 'Total', 'Matches', 'Shots', 'Accuracy', 'Tackles']
    for col, header in enumerate(player_headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment if col in [1, 7] else left_alignment

    # Player stats rows
    for player in sorted(player_stats.values(), key=lambda x: x['number'] or 999):
        player_acc = (player['shots_on_target'] / player['shots_taken'] * 100) if player['shots_taken'] > 0 else 0
        player_total = player['goals'] * 3 + player['point_1'] + player['point_2']

        cells = [
            player['number'] or '-',
            player['name'],
            player['goals'],
            player['point_1'],
            player['point_2'],
            player_total,
            len(player['matches']),
            player['shots_taken'],
            f"{player_acc:.1f}%",
            player['tackles_won'],
        ]

        for col, value in enumerate(cells, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.alignment = center_alignment if col in [1, 7] else left_alignment

        row += 1

    # Column widths
    ws.column_dimensions['A'] = 12
    ws.column_dimensions['B'] = 30
    ws.column_dimensions['C'] = 10
    ws.column_dimensions['D'] = 12
    ws.column_dimensions['E'] = 10
    ws.column_dimensions['F'] = 12
    ws.column_dimensions['G'] = 12
    ws.column_dimensions['H'] = 12
    ws.column_dimensions['I'] = 12
    ws.column_dimensions['J'] = 12
    ws.row_dimensions[1] = 30
    ws.row_dimensions[row-2] = 30

    output_path = REPORTS_DIR / 'season_report.xlsx'
    wb.save(output_path)

    return output_path
