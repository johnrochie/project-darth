"""
Player Report Generation
PDF and Excel reports for individual players
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from django.conf import settings
from ..models import Match, Player, MatchEvent, MatchParticipant

REPORTS_DIR = settings.MEDIA_ROOT / 'reports'
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def generate_player_report_pdf(player: Player) -> str:
    """Generate PDF report for a specific player"""

    # Get all participations
    participations = MatchParticipant.objects.filter(player=player).select_related('match')

    # Get all events for this player
    events = MatchEvent.objects.filter(player=player).select_related('match').order_by('-timestamp')[:50]

    # Calculate stats
    stats = {
        'matches_played': participations.count(),
        'goals': 0,
        'point_1': 0,
        'point_2': 0,
        'shots_taken': 0,
        'shots_on_target': 0,
        'tackles_won': 0,
    }

    for event in events:
        if event.event_type == 'score_goal':
            stats['goals'] += 1
        elif event.event_type == 'score_1point':
            stats['point_1'] += 1
        elif event.event_type == 'score_2point':
            stats['point_2'] += 1
        elif event.event_type == 'shot_on_target':
            stats['shots_taken'] += 1
            stats['shots_on_target'] += 1
        elif event.event_type == 'shot_wide':
            stats['shots_taken'] += 1
        elif event.event_type == 'shot_saved':
            stats['shots_taken'] += 1
        elif event.event_type == 'tackle_won':
            stats['tackles_won'] += 1

    total_points = stats['goals'] * 3 + stats['point_1'] + stats['point_2']
    shot_accuracy = (stats['shots_on_target'] / stats['shots_taken'] * 100) if stats['shots_taken'] > 0 else 0

    # HTML template
    html_string = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; }}
            .header {{ text-align: center; margin-bottom: 40px; }}
            .header h1 {{ color: #10B981; font-size: 28px; margin: 0; }}
            .header .subtitle {{ color: #6B7280; font-size: 16px; margin-top: 8px; }}
            .player-info {{ background: #F9FAFB; padding: 20px; border-radius: 8px; margin-bottom: 30px; }}
            .player-info h2 {{ margin: 0 0 8px 0; color: #1F2937; }}
            .info-grid {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 8px; }}
            .stat-display {{ display: flex; gap: 20px; margin-top: 20px; }}
            .stat-box {{ background: #10B981; color: white; padding: 20px; border-radius: 8px; text-align: center; }}
            .stat-box .value {{ font-size: 36px; font-weight: bold; font-size: 28px; margin: 0; }}
            .stat-box .label {{ font-size: 14px; opacity: 0.9; margin-top: 4px; }}
            .stats-table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            .stats-table th {{ background: #10B981; color: white; padding: 12px; text-align: left; font-weight: 600; }}
            .stats-table td {{ padding: 12px; border-bottom: 1px solid #E5E7EB; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>👤 Player Report</h1>
            <div class="subtitle">{player.name} - Season Summary</div>
        </div>

        <div class="player-info">
            <h2>Player Information</h2>
            <div class="info-grid">
                <div><strong>Number:</strong> {player.number or 'N/A'}</div>
                <div><strong>Position:</strong> {player.position or 'N/A'}</div>
                <div><strong>Status:</strong> {player.get_injury_status_display()}</div>
                <div><strong>Matches Played:</strong> {stats['matches_played']}</div>
            </div>
        </div>

        <div class="stat-display">
            <div class="stat-box">
                <div class="value">{total_points}</div>
                <div class="label">Total Points</div>
            </div>
            <div class="stat-box">
                <div class="value">{stats['goals']}</div>
                <div class="label">Goals</div>
            </div>
            <div class="stat-box">
                <div class="value">{stats['point_1']}</div>
                <div class="label">1-Points</div>
            </div>
            <div class="stat-box">
                <div class="value">{stats['point_2']}</div>
                <div class="label">2-Points</div>
            </div>
        </div>

        <h2>Statistics Breakdown</h2>
        <table class="stats-table">
            <tr>
                <th>Goals</th>
                <th>1-Points</th>
                <th>2-Points</th>
                <th>Total</th>
                <th>Shots</th>
                <th>Accuracy</th>
                <th>Tackles</th>
            </tr>
            <tr>
                <td>{stats['goals']}</td>
                <td>{stats['point_1']}</td>
                <td>{stats['point_2']}</td>
                <td>{total_points}</td>
                <td>{stats['shots_taken']}</td>
                <td>{shot_accuracy:.1f}%</td>
                <td>{stats['tackles_won']}</td>
            </tr>
        </table>

        <div style="text-align: center; margin-top: 60px; color: #9CA3AF; font-size: 12px;">
            Generated: {{datetime.now().strftime('%d %B %Y at %H:%M')}} | GAA Stats App
        </div>
    </body>
    </html>
    """

    from datetime import datetime
    from weasyprint import HTML

    output_path = REPORTS_DIR / f'player_{player.id}_report.pdf'
    HTML(string=html_string).write_pdf(target=str(output_path))

    return output_path


def generate_player_report_excel(player: Player) -> str:
    """Generate Excel report for a specific player"""

    participations = MatchParticipant.objects.filter(player=player).select_related('match')
    events = MatchEvent.objects.filter(player=player).select_related('match')

    wb = Workbook()
    ws = wb.active
    ws.title = f"{player.name} - Stats"

    # Styles
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center")
    left_alignment = Alignment(horizontal="left")

    # Player Info
    row = 1
    ws[f'A{row}'] = 'Player Report'
    ws[f'A{row}'].font = Font(bold=True, size=16)

    row += 2
    ws[f'A{row}'] = 'Player:'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = player.name

    row += 1
    ws[f'A{row}'] = 'Position:'
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = player.position or 'N/A'

    row += 2

    # Stats Header
    ws[f'A{row}'] = 'Season Statistics'
    ws[f'A{row}'].font = Font(bold=True)
    row += 1

    headers = ['Matches Played', 'Goals', '1-Points', '2-Points', 'Total', 'Shots', 'Accuracy', 'Tackles']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # Calculate stats
    stats = {
        'matches_played': participations.count(),
        'goals': 0,
        'point_1': 0,
        'point_2': 0,
        'shots_taken': 0,
        'shots_on_target': 0,
        'tackles_won': 0,
    }

    for event in events:
        if event.event_type == 'score_goal':
            stats['goals'] += 1
        elif event.event_type == 'score_1point':
            stats['point_1'] += 1
        elif event.event_type == 'score_2point':
            stats['point_2'] += 1
        elif event.event_type in ['shot_on_target', 'shot_wide', 'shot_saved']:
            stats['shots_taken'] += 1
            if event.event_type == 'shot_on_target':
                stats['shots_on_target'] += 1
        elif event.event_type == 'tackle_won':
            stats['tackles_won'] += 1

    total_points = stats['goals'] * 3 + stats['point_1'] + stats['point_2']
    accuracy = (stats['shots_on_target'] / stats['shots_taken'] * 100) if stats['shots_taken'] > 0 else 0

    row += 1
    values = [
        stats['matches_played'],
        stats['goals'],
        stats['point_1'],
        stats['point_2'],
        total_points,
        stats['shots_taken'],
        f"{accuracy:.1f}%",
        stats['tackles_won'],
    ]

    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.alignment = center_alignment

    # Column widths
    ws.column_dimensions['A'] = 20
    ws.column_dimensions['B'] = 12
    ws.column_dimensions['C'] = 12
    ws.column_dimensions['D'] = 12
    ws.column_dimensions['E'] = 12
    ws.column_dimensions['F'] = 12
    ws.column_dimensions['G'] = 12
    ws.column_dimensions['H'] = 12

    output_path = REPORTS_DIR / f'player_{player.id}_report.xlsx'
    wb.save(output_path)

    return output_path
