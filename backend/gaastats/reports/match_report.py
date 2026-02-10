"""
PDF and Excel Report Generation
"""

from datetime import datetime
import os
from weasyprint import HTML, CSS
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.conf import settings
from ..models import Match, Player, MatchEvent

# Report output directory
REPORTS_DIR = settings.MEDIA_ROOT / 'reports'
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def generate_match_report_pdf(match: Match) -> str:
    """Generate PDF report for a specific match"""

    # Get events
    events = MatchEvent.objects.filter(
        match=match
    ).select_related('player').order_by('minute')

    # Calculate team stats
    team_stats = {
        'goals': 0,
        'point_1': 0,
        'point_2': 0,
        'shots_taken': 0,
        'shots_on_target': 0,
        'tackles_won': 0,
    }

    player_stats = {}
    for event in events:
        if event.event_type == 'score_goal':
            team_stats['goals'] += 1
        elif event.event_type == 'score_1point':
            team_stats['point_1'] += 1
        elif event.event_type == 'score_2point':
            team_stats['point_2'] += 1
        elif event.event_type == 'shot_on_target':
            team_stats['shots_taken'] += 1
            team_stats['shots_on_target'] += 1
        elif event.event_type == 'shot_wide':
            team_stats['shots_taken'] += 1
        elif event.event_type == 'shot_saved':
            team_stats['shots_taken'] += 1
        elif event.event_type == 'tackle_won':
            team_stats['tackles_won'] += 1

        # Player-specific stats
        if event.player:
            if event.player_id not in player_stats:
                player_stats[event.player_id] = {
                    'name': event.player.name,
                    'number': event.player.number,
                    'goals': 0,
                    'point_1': 0,
                    'point_2': 0,
                    'shots_taken': 0,
                    'shots_on_target': 0,
                    'tackles_won': 0,
                }

            stats = player_stats[event.player_id]

            if event.event_type == 'score_goal':
                stats['goals'] += 1
            elif event.event_type == 'score_1point':
                stats['point_1'] += 1
            elif event.event_type == 'score_2point':
                stats['point_2'] += 1
            elif event.event_type == 'shot_on_target':
                stats['shots_taken'] += 1
                stats['shots_on_target'] += 1
            elif event.event_type == 'shot_wide':
                stats['shots_taken'] += 1
            elif event.event_type == 'shot_saved':
                stats['shots_taken'] += 1
            elif event.event_type == 'tackle_won':
                stats['tackles_won'] += 1

    total_score = team_stats['goals'] * 3 + team_stats['point_1'] + team_stats['point_2']
    shot_accuracy = (
        (team_stats['shots_on_target'] / team_stats['shots_taken'] * 100)
        if team_stats['shots_taken'] > 0
        else 0
    )

    # HTML template
    html_string = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; }}
            .header {{ text-align: center; margin-bottom: 40px; }}
            .header h1 {{ color: #10B981; font-size: 28px; margin: 0; }}
            .header .subtitle {{ color: #6B7280; font-size: 16px; margin-top: 8px; }}
            .match-info {{ background: #F9FAFB; padding: 20px; border-radius: 8px; margin-bottom: 30px; }}
            .match-info h2 {{ margin: 0 0 16px 0; color: #1F2937; }}
            .info-row {{ display: flex; justify-content: space-between; margin-bottom: 8px; }}
            .info-label {{ color: #6B7280; font-weight: 600; }}
            .info-value {{ color: #1F2937; font-weight: 500; }}
            .score-display {{ text-align: center; background: #10B981; color: white; padding: 24px; border-radius: 12px; margin-bottom: 30px; }}
            .score-display h3 {{ margin: 0 0 8px 0; font-size: 18px; opacity: 0.9; }}
            .score-display .score {{ font-size: 56px; font-weight: bold; margin: 0; }}
            .score-display .score-breakdown {{ font-size: 18px; opacity: 0.9; }}
            .stats-table {{ width: 100%; border-collapse: collapse; margin-bottom: 30px; }}
            .stats-table th {{ background: #10B981; color: white; padding: 12px; text-align: left; font-weight: 600; }}
            .stats-table td {{ padding: 12px; border-bottom: 1px solid #E5E7EB; }}
            .stats-table tr:last-child td {{ border-bottom: none; }}
            .player-section {{ margin-top: 40px; }}
            .player-section h2 {{ color: #1F2937; margin-bottom: 20px; }}
            .player-table {{ width: 100%; border-collapse: collapse; margin-top: 16px; }}
            .player-table th {{ background: #F3F4F6; color: #1F2937; padding: 12px; text-align: left; font-weight: 600; }}
            .player-table td {{ padding: 10px; border-bottom: 1px solid #E5E7EB; text-align: center; }}
            .player-table .number {{ font-weight: bold; color: #10B981; }}
            .player-table .name {{ text-align: left; font-weight: 500; }}
            .player-table tr:last-child td {{ border-bottom: none; }}
        </style>
    </head>
    <body>
        <div class="header">
            <h1>🏈 GAA Match Report</h1>
            <div class="subtitle">{match.club.name} - Match Summary</div>
        </div>

        <div class="match-info">
            <h2>Match Information</h2>
            <div class="info-row">
                <span class="info-label">Opposition:</span>
                <span class="info-value">{match.opposition}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Date:</span>
                <span class="info-value">{match.date.strftime('%A, %d %B %Y')}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Venue:</span>
                <span class="info-value">{match.venue or 'TBD'}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Weather:</span>
                <span class="info-value">{match.weather or 'N/A'}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Referee:</span>
                <span class="info-value">{match.referee or 'N/A'}</span>
            </div>
            <div class="info-row">
                <span class="info-label">Competition:</span>
                <span class="info-value">{match.competition or 'N/A'}</span>
            </div>
        </div>

        <div class="score-display">
            <h3>Final Score</h3>
            <div class="score">{total_score} - {match.total_opposition_score}</div>
            <div class="score-breakdown">{match.club_goals}-{match.club_1point}-{match.club_2point} vs {match.opposition_goals}-0-0</div>
        </div>

        <h2>Team Statistics</h2>
        <table class="stats-table">
            <tr>
                <th>Goals</th>
                <th>1-Points</th>
                <th>2-Points</th>
                <th>Total</th>
                <th>Shots</th>
                <th>Accuracy</th>
                <th>Tackles</th>
            </tr>
            <tr>
                <td>{team_stats['goals']}</td>
                <td>{team_stats['point_1']}</td>
                <td>{team_stats['point_2']}</td>
                <td>{total_score}</td>
                <td>{team_stats['shots_taken']}</td>
                <td>{shot_accuracy:.1f}%</td>
                <td>{team_stats['tackles_won']}</td>
            </tr>
        </table>

        <div class="player-section">
            <h2>Player Statistics</h2>
            <table class="player-table">
                <tr>
                    <th>#</th>
                    <th class="name">Player</th>
                    <th>Goals</th>
                    <th>Pts</th>
                    <th>2-Pts</th>
                    <th>Shots</th>
                    <th>Acc</th>
                    <th>Tackles</th>
                </tr>
                {"".join(f"""
                <tr>
                    <td class="number">{p['number'] or '-'}</td>
                    <td class="name">{p['name']}</td>
                    <td>{p['goals']}</td>
                    <td>{p['point_1']}</td>
                    <td>{p['point_2']}</td>
                    <td>{p['shots_taken']}</td>
                    <td>{(p['shots_on_target'] / p['shots_taken'] * 100) if p['shots_taken'] > 0 else 0:.1f}%</td>
                    <td>{p['tackles_won']}</td>
                </tr>
                """ for p in sorted(player_stats.values(), key=lambda x: x['number'] or 999))}
            </table>
        </div>

        <div style="text-align: center; margin-top: 60px; color: #9CA3AF; font-size: 12px;">
            Generated: {datetime.now().strftime('%d %B %Y at %H:%M')} | GAA Stats App
        </div>
    </body>
    </html>
    """

    # Generate PDF
    output_path = REPORTS_DIR / f'match_{match.id}_report.pdf'
    HTML(string=html_string).write_pdf(target=str(output_path))

    return output_path


def generate_match_report_excel(match: Match) -> str:
    """Generate Excel report for a specific match"""

    events = MatchEvent.objects.filter(match=match).select_related('player').order_by('minute')

    wb = Workbook()
    ws = wb.active
    ws.title = f"Match Report - {match.opposition}"

    # Styles
    header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    center_alignment = Alignment(horizontal="center")
    left_alignment = Alignment(horizontal="left")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Match Info
    row = 1
    ws[f'A{row}'] = 'Match Report'
    ws[f'A{row}'].font = Font(bold=True, size=16)

    row += 2
    ws[f'A{row}'] = 'Opposition:'
    ws[f'B{row}'] = match.opposition
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = 'Date:'
    ws[f'B{row}'] = match.date.strftime('%Y-%m-%d')

    row += 1
    ws[f'A{row}'] = 'Venue:'
    ws[f'B{row}'] = match.venue or 'TBD'

    row += 2

    # Team Stats Header
    ws[f'A{row}'] = 'Team Statistics'
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    headers = ['Goals', '1-Points', '2-Points', 'Total', 'Shots', 'Accuracy', 'Tackles']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment

    # Calculate stats (same as PDF logic)
    team_stats = {'goals': 0, 'point_1': 0, 'point_2': 0, 'shots_taken': 0, 'shots_on_target': 0, 'tackles_won': 0}
    for event in events:
        if event.event_type == 'score_goal':
            team_stats['goals'] += 1
        elif event.event_type == 'score_1point':
            team_stats['point_1'] += 1
        elif event.event_type == 'score_2point':
            team_stats['point_2'] += 1
        elif event.event_type in ['shot_on_target', 'shot_wide', 'shot_saved']:
            team_stats['shots_taken'] += 1
            if event.event_type == 'shot_on_target':
                team_stats['shots_on_target'] += 1
        elif event.event_type == 'tackle_won':
            team_stats['tackles_won'] += 1

    total_score = team_stats['goals'] * 3 + team_stats['point_1'] + team_stats['point_2']
    accuracy = (team_stats['shots_on_target'] / team_stats['shots_taken'] * 100) if team_stats['shots_taken'] > 0 else 0

    row += 1
    stats = [team_stats['goals'], team_stats['point_1'], team_stats['point_2'], total_score, team_stats['shots_taken'], f"{accuracy:.1f}%", team_stats['tackles_won']]
    for col, value in enumerate(stats, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = value
        cell.alignment = center_alignment

    row += 2

    # Player Stats Header
    ws[f'A{row}'] = 'Player Statistics'
    ws[f'A{row}'].font = Font(bold=True, size=14)
    row += 1

    player_headers = ['#', 'Player', 'Goals', '1-Points', '2-Points', 'Shots', 'Accuracy', 'Tackles']
    for col, header in enumerate(player_headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment if col in [1, 7] else left_alignment

    # Player stats
    player_stats = {}
    for event in events:
        if event.player:
            if event.player_id not in player_stats:
                player_stats[event.player_id] = {
                    'number': event.player.number,
                    'name': event.player.name,
                    'goals': 0,
                    'point_1': 0,
                    'point_2': 0,
                    'shots_taken': 0,
                    'shots_on_target': 0,
                    'tackles_won': 0,
                }

            stats = player_stats[event.player_id]

            if event.event_type == 'score_goal':
                stats['goals'] += 1
            elif event.event_type == 'score_1point':
                stats['point_1'] += 1
            elif event.event_type == 'score_2point':
                stats['point_2'] += 1
            elif event.event_type in ['shot_on_target', 'shot_wide', 'shot_saved']:
                stats['shots_taken'] += 1
                if event.event_type == 'shot_on_target':
                    stats['shots_on_target'] += 1
            elif event.event_type == 'tackle_won':
                stats['tackles_won'] += 1

    row += 1
    for player in sorted(player_stats.values(), key=lambda x: x['number'] or 999):
        player_acc = (player['shots_on_target'] / player['shots_taken'] * 100) if player['shots_taken'] > 0 else 0

        cells = [
            player['number'] or '-',
            player['name'],
            player['goals'],
            player['point_1'],
            player['point_2'],
            player['shots_taken'],
            f"{player_acc:.1f}%",
            player['tackles_won'],
        ]

        for col, value in enumerate(cells, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.alignment = center_alignment if col in [1, 7] else left_alignment

        row += 1

    # Column widths
    ws.column_dimensions['A'] = 15
    ws.column_dimensions['B'] = 30
    ws.column_dimensions['C'] = 10
    ws.column_dimensions['D'] = 10
    ws.column_dimensions['E'] = 10
    ws.column_dimensions['F'] = 10
    ws.column_dimensions['G'] = 10
    ws.column_dimensions['H'] = 12

    output_path = REPORTS_DIR / f'match_{match.id}_report.xlsx'
    wb.save(output_path)

    return output_path
